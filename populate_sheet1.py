#!/usr/bin/env python3
"""Populate Sheet1 pricing and distance fields using data from other sheets."""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

try:
    import googlemaps
except ImportError as exc:  # pragma: no cover - handled at runtime
    googlemaps = None  # type: ignore


Number = Optional[float]


@dataclass
class BranchZone:
    service_level: str
    zone_label: str
    min_miles: float
    max_miles: float
    branch_delivery_charge: Number
    driver_rate: Number
    branch_over_mile_rate: Number
    driver_over_mile_rate: Number


@dataclass
class DriverHelperZone:
    zone_label: str
    min_miles: float
    max_miles: float
    driver_rate: Number
    helper_rate: Number


class DistanceProvider:
    """Wrap Google Maps distance lookups with caching."""

    def __init__(self, api_key: str) -> None:
        if googlemaps is None:
            raise RuntimeError(
                "googlemaps package not available. Install it or adjust PYTHONPATH."
            )
        if not api_key:
            raise ValueError("Google Maps API key must be provided.")
        self._client = googlemaps.Client(key=api_key)
        self._cache: Dict[Tuple[str, str], float] = {}

    def get_distance_miles(self, origin_zip: str, destination_zip: str) -> float:
        key = (origin_zip, destination_zip)
        if key in self._cache:
            return self._cache[key]

        response = self._client.distance_matrix(
            origins=[origin_zip],
            destinations=[destination_zip],
            mode="driving",
            units="imperial",
        )

        try:
            element = response["rows"][0]["elements"][0]
        except (KeyError, IndexError) as err:
            raise RuntimeError(f"Unexpected Google Maps response: {response}") from err

        status = element.get("status")
        if status != "OK":
            raise RuntimeError(
                f"Distance lookup failed for {origin_zip}->{destination_zip}: {status}"
            )

        distance_text = element.get("distance", {}).get("text")
        distance_value = element.get("distance", {}).get("value")
        if distance_value is None:
            raise RuntimeError(
                f"Distance value missing in element for {origin_zip}->{destination_zip}: {element}"
            )

        # Distance returned in meters; convert to miles.
        miles = float(distance_value) / 1609.344

        # Guard against locales returning text in km even with imperial.
        if distance_text and "mi" in distance_text:
            numeric_text = re.sub(r"[^0-9.]+", "", distance_text)
            try:
                text_value = float(numeric_text)
            except ValueError:
                text_value = miles
            else:
                # Align conversion with displayed miles if they differ slightly.
                miles = text_value

        self._cache[key] = miles
        return miles


def normalize_header(value: str) -> str:
    """Normalize header labels for more reliable matching."""

    return re.sub(r"\s+", " ", value).strip().upper()


def find_header_row(ws: Worksheet, target_header: str) -> int:
    target = normalize_header(target_header)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value and normalize_header(str(cell.value)) == target:
                return cell.row
    raise ValueError(f"Could not find header '{target_header}' in sheet '{ws.title}'.")


def build_header_lookup(ws: Worksheet, header_row: int) -> Dict[str, List[int]]:
    lookup: Dict[str, List[int]] = {}
    for cell in ws[header_row]:
        if not cell.value:
            continue
        header = normalize_header(str(cell.value))
        lookup.setdefault(header, []).append(cell.column)
    return lookup


def parse_zone_range(text: str) -> Tuple[float, float]:
    match = re.search(r"(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)", text)
    if not match:
        raise ValueError(f"Could not parse zone range from '{text}'.")
    return float(match.group(1)), float(match.group(2))


def parse_money(value: Optional[str | float | int]) -> Number:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = value.replace("$", "").replace(",", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_over_mile_rate(text: Optional[str]) -> Number:
    if not text:
        return None
    match = re.search(r"\$(\d+(?:\.\d+)?)\s*/?\s*Mile", text, re.IGNORECASE)
    if match:
        return float(match.group(1))
    return None


def load_branch_zones(ws: Worksheet) -> Dict[str, List[BranchZone]]:
    zones_by_level: Dict[str, List[BranchZone]] = {}
    current_level: Optional[str] = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        service_level, zone_label, zone_range, branch_charge, driver_rate, _ = row[:6]

        if service_level and normalize_header(str(service_level)) == "ACCESSORIAL RATES":
            break

        if service_level:
            current_level = normalize_header(str(service_level))
            zones_by_level.setdefault(current_level, [])

        if not current_level or not zone_label or not zone_range:
            continue

        zone_label_norm = normalize_header(str(zone_label))
        if not isinstance(zone_range, str):
            continue

        min_miles, max_miles = parse_zone_range(zone_range)
        branch_charge_value = parse_money(branch_charge)
        driver_rate_value = parse_money(driver_rate)
        branch_over_rate = parse_over_mile_rate(branch_charge if isinstance(branch_charge, str) else "")
        driver_over_rate = parse_over_mile_rate(driver_rate if isinstance(driver_rate, str) else "")

        zone = BranchZone(
            service_level=current_level,
            zone_label=zone_label_norm,
            min_miles=min_miles,
            max_miles=max_miles,
            branch_delivery_charge=branch_charge_value,
            driver_rate=driver_rate_value,
            branch_over_mile_rate=branch_over_rate,
            driver_over_mile_rate=driver_over_rate,
        )
        zones_by_level[current_level].append(zone)

    return zones_by_level


def load_extra_piece_rates(ws: Worksheet) -> Tuple[Number, Number]:
    for row in ws.iter_rows(values_only=True):
        first_cell = row[0]
        if isinstance(first_cell, str) and normalize_header(first_cell) == "MORE THAN 4 PCS":
            branch_rate = parse_money(row[2])
            driver_rate = parse_money(row[3])
            return branch_rate, driver_rate
    raise ValueError("Could not locate 'More than 4 pcs' accessorial rates in Sheet2.")


def load_origin_zip(ws: Worksheet) -> str:
    for row in ws.iter_rows(values_only=True):
        first_cell = row[0]
        if isinstance(first_cell, str) and normalize_header(first_cell).startswith(
            "ORIGINAL ZIP CODE"
        ):
            zip_value = row[1]
            if isinstance(zip_value, (int, float)):
                return f"{int(zip_value):05d}"
            if isinstance(zip_value, str):
                digits = re.sub(r"\D", "", zip_value)
                if digits:
                    return digits.zfill(5)
    raise ValueError("Original zip code not found in Sheet2.")


def load_driver_helper_zones(ws: Worksheet) -> List[DriverHelperZone]:
    zones: List[DriverHelperZone] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        zone_label = row[2]
        if not zone_label:
            continue
        zone_label_norm = normalize_header(str(zone_label))
        if not zone_label_norm.startswith("ZONE"):
            continue
        zone_range = row[3]
        if not isinstance(zone_range, str):
            continue
        min_miles, max_miles = parse_zone_range(zone_range)
        driver_rate = parse_money(row[4])
        helper_rate = parse_money(row[5])
        zones.append(
            DriverHelperZone(
                zone_label=zone_label_norm,
                min_miles=min_miles,
                max_miles=max_miles,
                driver_rate=driver_rate,
                helper_rate=helper_rate,
            )
        )

    if not zones:
        raise ValueError("No driver/helper zones found in Sheet3.")

    return zones


def find_branch_zone(
    service_level: str, distance: float, zones_by_level: Dict[str, List[BranchZone]]
) -> BranchZone:
    normalized = normalize_header(service_level)
    if normalized not in zones_by_level:
        raise ValueError(
            f"Service level '{service_level}' not found in Sheet2. Available: {list(zones_by_level)}"
        )

    for zone in zones_by_level[normalized]:
        if zone.min_miles <= distance <= zone.max_miles:
            return zone

    # If distance exceeds configured max, fall back to highest zone.
    return zones_by_level[normalized][-1]


def find_driver_helper_zone(distance: float, zones: List[DriverHelperZone]) -> DriverHelperZone:
    for zone in zones:
        if zone.min_miles <= distance <= zone.max_miles:
            return zone
    return zones[-1]


def compute_branch_charges(
    zone: BranchZone, zones_by_level: Dict[str, List[BranchZone]], distance: float
) -> Tuple[float, float]:
    over_pay = 0.0
    branch_total = zone.branch_delivery_charge or 0.0

    if zone.branch_delivery_charge is not None:
        branch_total = zone.branch_delivery_charge
    elif zone.branch_over_mile_rate is not None:
        # Use preceding zone (Zone 4) as base.
        base_zone = _find_previous_zone(zone, zones_by_level)
        base_charge = base_zone.branch_delivery_charge or 0.0
        over_miles = max(distance - base_zone.max_miles, 0.0)
        over_pay = over_miles * zone.branch_over_mile_rate
        branch_total = base_charge + over_pay

    return branch_total, over_pay


def compute_driver_charges(
    zone: BranchZone, zones_by_level: Dict[str, List[BranchZone]], distance: float
) -> float:
    if zone.driver_rate is not None:
        return zone.driver_rate

    if zone.driver_over_mile_rate is None:
        return 0.0

    base_zone = _find_previous_zone(zone, zones_by_level)
    base_rate = base_zone.driver_rate or 0.0
    over_miles = max(distance - base_zone.max_miles, 0.0)
    return base_rate + over_miles * zone.driver_over_mile_rate


def _find_previous_zone(zone: BranchZone, zones_by_level: Dict[str, List[BranchZone]]) -> BranchZone:
    level_zones = zones_by_level[zone.service_level]
    idx = level_zones.index(zone)
    if idx == 0:
        return zone
    return level_zones[idx - 1]


def ensure_float(value: Number) -> Optional[float]:
    return None if value is None else float(value)


def format_zip(zip_value) -> Optional[str]:
    if zip_value is None:
        return None
    if isinstance(zip_value, (int, float)):
        return f"{int(zip_value):05d}"
    digits = re.sub(r"\D", "", str(zip_value))
    if not digits:
        return None
    return digits.zfill(5)


def populate_sheet(
    workbook_path: str,
    output_path: Optional[str],
    api_key: str,
    sheet1_name: str,
) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    ws1 = wb[sheet1_name]
    ws2 = wb["Sheet2"]
    ws3 = wb["Sheet3"]

    header_row = find_header_row(ws1, "VENDOR")
    headers = build_header_lookup(ws1, header_row)

    def require(header: str, index: int = 0) -> int:
        key = normalize_header(header)
        if key not in headers:
            raise ValueError(f"Column '{header}' not found in Sheet1 header row {header_row}.")
        try:
            return headers[key][index]
        except IndexError as err:
            raise ValueError(
                f"Column '{header}' occurrence {index} not found in Sheet1 header."
            ) from err

    col_zip = require("ZIP CODE")
    col_qty = require("QTY PIECE")
    col_service_level = require("BRANCH SERVICE LEVEL")
    col_branch_distance = require("DISTANCE", index=0)
    col_branch_zone = require("ZONE")
    col_branch_charge = require("BRANCH DELIVERY CHARGES")
    col_over_100_pay = require("OVER 100 MILES PAY")
    col_extra_pieces_branch = require("BRANCH RATE FOR EXTRA PIECES")
    col_driver_distance = require("DISTANCE", index=1)
    col_zone_type = require("ZONE TYPE")
    col_driver_rate = require("DRIVER RATE")
    col_extra_pieces_driver = require("EXTRA PIECES DRIVER RATE")
    col_helper_name = require("HELPER")
    col_helper_zone = require("HELPER ZONE 3, 4 & 5")
    col_helper_rate = require("HELPER RATE")

    origin_zip = load_origin_zip(ws2)
    branch_zones = load_branch_zones(ws2)
    branch_extra_rate, driver_extra_rate = load_extra_piece_rates(ws2)
    driver_helper_zones = load_driver_helper_zones(ws3)

    distance_provider = DistanceProvider(api_key=api_key)

    for row_idx in range(header_row + 1, ws1.max_row + 1):
        vendor_value = ws1.cell(row=row_idx, column=1).value
        if vendor_value is None:
            continue

        zip_code = format_zip(ws1.cell(row=row_idx, column=col_zip).value)
        if not zip_code:
            continue

        service_level_raw = ws1.cell(row=row_idx, column=col_service_level).value
        if not service_level_raw:
            continue
        service_level = str(service_level_raw)

        qty_value = ws1.cell(row=row_idx, column=col_qty).value
        qty = float(qty_value) if isinstance(qty_value, (int, float)) else 0.0

        distance = distance_provider.get_distance_miles(origin_zip, zip_code)
        distance = round(distance, 2)

        branch_zone = find_branch_zone(service_level, distance, branch_zones)
        driver_helper_zone = find_driver_helper_zone(distance, driver_helper_zones)

        branch_charge, over_100_pay = compute_branch_charges(branch_zone, branch_zones, distance)
        driver_rate_value = compute_driver_charges(branch_zone, branch_zones, distance)

        extra_piece_count = max(int(round(qty)) - 4, 0)
        branch_extra_total = (branch_extra_rate or 0.0) * extra_piece_count
        driver_extra_total = (driver_extra_rate or 0.0) * extra_piece_count

        helper_name = ws1.cell(row=row_idx, column=col_helper_name).value
        helper_zone_value = ""
        helper_rate_value: Optional[float] = None
        if helper_name:
            helper_zone_value = driver_helper_zone.zone_label.title()
            if branch_zone.zone_label.endswith(("3", "4", "5")):
                helper_rate_value = ensure_float(driver_helper_zone.helper_rate)
            else:
                helper_rate_value = 0.0

        ws1.cell(row=row_idx, column=col_branch_distance, value=distance)
        ws1.cell(row=row_idx, column=col_branch_zone, value=branch_zone.zone_label.title())
        ws1.cell(row=row_idx, column=col_branch_charge, value=round(branch_charge, 2))
        ws1.cell(row=row_idx, column=col_over_100_pay, value=round(over_100_pay, 2))
        ws1.cell(row=row_idx, column=col_extra_pieces_branch, value=round(branch_extra_total, 2))

        ws1.cell(row=row_idx, column=col_driver_distance, value=distance)
        ws1.cell(row=row_idx, column=col_zone_type, value=driver_helper_zone.zone_label.title())
        ws1.cell(row=row_idx, column=col_driver_rate, value=round(driver_rate_value, 2))
        ws1.cell(row=row_idx, column=col_extra_pieces_driver, value=round(driver_extra_total, 2))

        if helper_zone_value:
            ws1.cell(row=row_idx, column=col_helper_zone, value=helper_zone_value)
        if helper_rate_value is not None:
            ws1.cell(row=row_idx, column=col_helper_rate, value=round(helper_rate_value, 2))

    wb.save(output_path or workbook_path)


def parse_args(argv: Optional[Iterable[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Populate Sheet1 columns using Sheet2/Sheet3 data and Google Maps distance.",
    )
    parser.add_argument(
        "workbook",
        help="Path to the Excel workbook (will be overwritten unless --output provided).",
    )
    parser.add_argument(
        "--sheet",
        default="Sheet1",
        help="Name of Sheet1 tab to update (default: Sheet1).",
    )
    parser.add_argument(
        "--output",
        help="Optional path for the updated workbook. Defaults to in-place overwrite.",
    )
    parser.add_argument(
        "--api-key",
        dest="api_key",
        help="Google Maps API key. If omitted, the GOOG_API_KEY or GOOGLE_MAPS_API_KEY env vars are used.",
    )
    return parser.parse_args(argv)


def main(argv: Optional[Iterable[str]] = None) -> int:
    args = parse_args(argv)

    api_key = args.api_key or os.getenv("GOOGLE_MAPS_API_KEY") or os.getenv("GOOG_API_KEY")
    if not api_key:
        print(
            "ERROR: Google Maps API key is required. Pass via --api-key or set GOOGLE_MAPS_API_KEY.",
            file=sys.stderr,
        )
        return 1

    try:
        populate_sheet(
            workbook_path=args.workbook,
            output_path=args.output,
            api_key=api_key,
            sheet1_name=args.sheet,
        )
    except Exception as exc:  # pragma: no cover - runtime safeguard
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

